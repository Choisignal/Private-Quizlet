#셀 스타일 변경하기 1
from openpyxl.styles import Font
from openpyxl import load_workbook
data_direct = "./학습자료/단답형/"
filename = "영어_단어"
wb = load_workbook(f"{data_direct}{filename}.xlsx") # test.xlsx 파일을 가져온다
ws = wb.active # 활성화
'''
##번호, 영어, 수학
a1 = ws["A1"] #번호
b1 = ws["B1"] #영어
c1 = ws["C1"] #수학

#스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자색 빨간색, 이텔릭체, 두껍게
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 글자색 보라색, Arial, 취소선
c1.font = Font(color="0000FF", size=20, underline="single") #글자색 파란색, 글자크기20, 밑줄
'''
ws.font = Font(color="0000FF", size=20, underline="single")
wb.save("test_style.xlsx")